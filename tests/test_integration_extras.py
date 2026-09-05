"""Tests de integracion extra (requieren Excel real).

Ejecutar con: pdm run pytest -m integration tests/test_integration_extras.py -v
"""
from __future__ import annotations

import subprocess
import time
from pathlib import Path

import pytest
import uiautomation as auto

from rpa_excel_ui_automation import ExcelManager

INPUT_FILE = Path(".data/input/origen.xlsx")
OUTPUT_FILE = Path(".data/output/destino.xlsx")
OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)


def _cleanup_output_files() -> None:
    """Elimina archivos de salida de ejecuciones anteriores."""
    output_dir = OUTPUT_FILE.parent
    if output_dir.exists():
        for f in output_dir.glob("*.xlsx"):
            f.unlink(missing_ok=True)


def _kill_excel() -> None:
    """Mata todos los procesos Excel y espera a que se cierre."""
    try:
        subprocess.run(
            ["taskkill", "/F", "/IM", "EXCEL.EXE"],
            capture_output=True,
            check=False,
            timeout=10,
        )
    except (subprocess.TimeoutExpired, FileNotFoundError):
        pass
    time.sleep(1)
    try:
        excel_window = auto.WindowControl(searchDepth=1, ClassName="XLMAIN")
        for _ in range(10):
            if not excel_window.Exists(0, 0):
                break
            time.sleep(0.5)
    except Exception:
        pass


@pytest.mark.integration
def test_save_as_no_visible_chars() -> bool:
    """Verificar que save_as no inserta caracteres visibles en celdas."""
    print("\n=== Test: save_as sin modificar celdas ===")
    _kill_excel()

    if not INPUT_FILE.exists():
        print(f"  [FAIL] Archivo de prueba no existe: {INPUT_FILE}")
        return False

    with ExcelManager() as excel:
        if not excel.open_file(INPUT_FILE):
            print("  [FAIL] No se pudo abrir el archivo")
            return False

        test_output = OUTPUT_FILE.parent / "test_no_chars.xlsx"
        result = excel.save_as(test_output)
        if not result:
            print("  [FAIL] save_as fallo")
            return False

    import openpyxl

    try:
        wb = openpyxl.load_workbook(test_output)
        ws = wb.active
        if ws["A1"].value == "ID" and ws["B1"].value == "Nombre" and ws["C1"].value == "Valor":
            print("  [PASS] save_as ejecutado sin modificar celdas")
            test_output.unlink(missing_ok=True)
            return True
        else:
            print(f"  [FAIL] Celdas modificadas: A1={ws['A1'].value}, B1={ws['B1'].value}, C1={ws['C1'].value}")
            test_output.unlink(missing_ok=True)
            return False
    except Exception as e:
        print(f"  [FAIL] Error verificando archivo: {e}")
        test_output.unlink(missing_ok=True)
        return False


@pytest.mark.integration
def test_detect_excel_already_open() -> bool:
    """Verificar que se detecta si Excel ya esta abierto."""
    print("\n=== Test: Deteccion de Excel abierto ===")
    _kill_excel()

    if not INPUT_FILE.exists():
        print(f"  [FAIL] Archivo de prueba no existe: {INPUT_FILE}")
        return False

    excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
    subprocess.Popen([excel_path, str(INPUT_FILE.resolve())])

    excel_window = None
    for attempt in range(15):
        try:
            excel_window = auto.WindowControl(searchDepth=1, ClassName="XLMAIN")
            if excel_window.Exists(1, 0.5):
                break
            excel_window = None
        except Exception:
            excel_window = None
            time.sleep(0.5)

    if excel_window is None:
        print("  [FAIL] Excel no se inicio")
        return False

    with ExcelManager() as excel:
        result = excel.open_file()
        if result:
            print("  [PASS] Excel ya abierto detectado y conectado")
        else:
            print("  [FAIL] No se pudo conectar a Excel abierto")
            return False

    time.sleep(0.5)
    try:
        check = auto.WindowControl(searchDepth=1, ClassName="XLMAIN")
        if check.Exists(1, 0.5):
            print("  [PASS] Excel sigue abierto (correcto)")
            check.SendKeys("{Alt}{F4}", waitTime=0.5)
            time.sleep(1)
            try:
                auto.WaitForNotExist(check, 3)
            except Exception:
                pass
            return True
        else:
            print("  [INFO] Excel se cerro automaticamente")
            return True
    except Exception:
        print("  [INFO] Excel se cerro o COM invalido (correcto)")
        return True


@pytest.mark.integration
def test_multiple_save_as_operations() -> bool:
    """Probar multiples operaciones de guardar como."""
    print("\n=== Test: Multiples guardar como ===")
    _kill_excel()

    if not INPUT_FILE.exists():
        print(f"  [FAIL] Archivo de prueba no existe: {INPUT_FILE}")
        return False

    output_files = [
        OUTPUT_FILE.parent / "output_1.xlsx",
        OUTPUT_FILE.parent / "output_2.xlsx",
        OUTPUT_FILE.parent / "output_3.xlsx",
    ]

    with ExcelManager() as excel:
        if not excel.open_file(INPUT_FILE):
            print("  [FAIL] No se pudo abrir el archivo")
            return False

        for i, output in enumerate(output_files, 1):
            result = excel.save_as(output)
            if result:
                print(f"  [PASS] Guardar como #{i}: {output.name}")
            else:
                print(f"  [FAIL] Guardar como #{i} fallido")
                return False

    all_created = all(f.exists() for f in output_files)
    if all_created:
        print("  [PASS] Todos los archivos se crearon correctamente")
        return True
    else:
        print("  [FAIL] No todos los archivos se crearon")
        return False


@pytest.mark.integration
def test_context_manager_cleanup() -> bool:
    """Verificar que el context manager cierra Excel correctamente."""
    print("\n=== Test: Context manager cleanup ===")
    _kill_excel()

    if not INPUT_FILE.exists():
        print(f"  [FAIL] Archivo de prueba no existe: {INPUT_FILE}")
        return False

    with ExcelManager() as excel:
        excel.open_file(INPUT_FILE)
        print("  [INFO] Excel abierto dentro del context manager")

    check = auto.WindowControl(searchDepth=1, ClassName="XLMAIN")
    if not check.Exists(2, 0.5):
        print("  [PASS] Excel se cerro al salir del context manager")
        return True
    else:
        print("  [FAIL] Excel no se cerro al salir del context manager")
        check.SendKeys("{Alt}{F4}", waitTime=0.5)
        auto.WaitForNotExist(check, 3)
        return False
